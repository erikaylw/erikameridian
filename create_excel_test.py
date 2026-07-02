import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

wb = openpyxl.Workbook()

# ======== STYLES ========
header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
subheader_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
subheader_font = Font(name='Calibri', bold=True, size=11)
normal_font = Font(name='Calibri', size=11)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
rupiah_fmt = '#,##0'
date_fmt = 'DD/MM/YYYY'

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

def style_range(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.font = normal_font

# ============================================================
# SHEET 1: Transaksi (Soal 1-4)
# ============================================================
ws1 = wb.active
ws1.title = 'Transaksi'

# Soal 1 headers
headers1 = ['Tanggal', 'Keterangan', 'Debit', 'Kredit', 'Saldo']
for c, h in enumerate(headers1, 1):
    ws1.cell(row=1, column=c, value=h)
style_header_row(ws1, 1, 5)

data = [
    ('01/06/2024', 'Saldo Awal', None, None, 10000000),
    ('03/06/2024', 'Pembelian ATK', 2000000, None, None),
    ('05/06/2024', 'Pendapatan Jasa', None, 5000000, None),
    ('07/06/2024', 'Bayar Listrik', 500000, None, None),
    ('10/06/2024', 'Bayar Gaji', 15000000, None, None),
]

for i, (tgl, ket, deb, kre, sal) in enumerate(data, 2):
    ws1.cell(row=i, column=1, value=tgl)
    ws1.cell(row=i, column=1).number_format = date_fmt
    ws1.cell(row=i, column=2, value=ket)
    if deb is not None:
        ws1.cell(row=i, column=3, value=deb)
        ws1.cell(row=i, column=3).number_format = rupiah_fmt
    if kre is not None:
        ws1.cell(row=i, column=4, value=kre)
        ws1.cell(row=i, column=4).number_format = rupiah_fmt
    if sal is not None:
        ws1.cell(row=i, column=5, value=sal)
        ws1.cell(row=i, column=5).number_format = rupiah_fmt

# Add hint text for Saldo column (row 2)
ws1.cell(row=2, column=6, value='Isi saldo awal')
ws1.cell(row=2, column=6).font = Font(name='Calibri', size=9, italic=True, color='888888')
ws1.cell(row=3, column=6, value='Rumus: saldo sblm + (kredit - debit)')
ws1.cell(row=3, column=6).font = Font(name='Calibri', size=9, italic=True, color='888888')

style_range(ws1, 2, 6, 1, 5)

# Soal 3: Kode & Status columns
ws1.cell(row=1, column=6, value='Kode')
ws1.cell(row=1, column=7, value='Status')
ws1.cell(row=1, column=8, value='Jenis Transaksi')
style_header_row(ws1, 1, 8)

# Kode data for soal 4
kode_data = ['ATK', 'JAS', 'LIS', 'GAJ']
for i, k in enumerate(kode_data, 3):
    ws1.cell(row=i, column=6, value=k)
    ws1.cell(row=i, column=6).font = normal_font
    ws1.cell(row=i, column=6).border = thin_border

# Blank Status column (Soal 3 - to be filled by candidate)
for i in range(3, 7):
    ws1.cell(row=i, column=7, value='')
    ws1.cell(row=i, column=7).border = thin_border

# Blank Jenis Transaksi column (Soal 7 - data validation)
for i in range(3, 13):
    ws1.cell(row=i, column=8, value='')
    ws1.cell(row=i, column=8).border = thin_border

# Soal 2: Summary area
summary_row = 9
ws1.cell(row=summary_row, column=1, value='RINGKASAN (Soal 2)')
ws1.cell(row=summary_row, column=1).font = Font(name='Calibri', bold=True, size=12, color='2F5496')
summaries = [
    ('Total Debit', ''),
    ('Total Kredit', ''),
    ('Rata-rata Debit', ''),
    ('Jumlah Transaksi', ''),
]
for i, (label, _) in enumerate(summaries, 1):
    r = summary_row + i
    ws1.cell(row=r, column=1, value=label)
    ws1.cell(row=r, column=1).font = subheader_font
    ws1.cell(row=r, column=2, value='')
    ws1.cell(row=r, column=2).border = thin_border
    ws1.cell(row=r, column=2).number_format = rupiah_fmt

# Column widths
ws1.column_dimensions['A'].width = 14
ws1.column_dimensions['B'].width = 20
ws1.column_dimensions['C'].width = 16
ws1.column_dimensions['D'].width = 16
ws1.column_dimensions['E'].width = 16
ws1.column_dimensions['F'].width = 10
ws1.column_dimensions['G'].width = 14
ws1.column_dimensions['H'].width = 18

# Conditional formatting hint
ws1.cell(row=summary_row + 6, column=1, value='Petunjuk:')
ws1.cell(row=summary_row + 6, column=1).font = Font(name='Calibri', bold=True, size=9, color='888888')
ws1.cell(row=summary_row + 7, column=1, value='SUM, AVERAGE, COUNT di kolom B')
ws1.cell(row=summary_row + 7, column=1).font = Font(name='Calibri', size=9, italic=True, color='888888')

# Data Validation for Jenis Transaksi (Soal 7)
dv = DataValidation(type='list', formula1='"Kas,Bank,Piutang,Utang"', allow_blank=True)
dv.error = 'Pilih dari dropdown'
dv.errorTitle = 'Input Error'
dv.prompt = 'Pilih jenis transaksi'
dv.promptTitle = 'Jenis Transaksi'
ws1.add_data_validation(dv)
for i in range(3, 13):
    dv.add(ws1.cell(row=i, column=8))

# Conditional Formatting for Debit (Soal 6)
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
white_font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')

ws1.conditional_formatting.add('C3:C12',
    CellIsRule(operator='greaterThan', formula=['5000000'], fill=red_fill, font=white_font))
ws1.conditional_formatting.add('C3:C12',
    CellIsRule(operator='between', formula=['500000', '5000000'], fill=yellow_fill))
ws1.conditional_formatting.add('C3:C12',
    CellIsRule(operator='lessThan', formula=['500000'], fill=green_fill))
# exclude zero/empty
ws1.conditional_formatting.add('C3:C12',
    CellIsRule(operator='equal', formula=['0'], fill=PatternFill()))  # no fill for zero

# ============================================================
# SHEET 2: Kategori (Soal 4 - VLOOKUP)
# ============================================================
ws2 = wb.create_sheet('Kategori')
ws2.cell(row=1, column=1, value='Kode')
ws2.cell(row=1, column=2, value='Kategori')
style_header_row(ws2, 1, 2)

kategori_data = [
    ('ATK', 'Perlengkapan Kantor'),
    ('JAS', 'Pendapatan Jasa'),
    ('LIS', 'Utilitas'),
    ('GAJ', 'Gaji Karyawan'),
]
for i, (k, v) in enumerate(kategori_data, 2):
    ws2.cell(row=i, column=1, value=k)
    ws2.cell(row=i, column=2, value=v)
    ws2.cell(row=i, column=1).font = normal_font
    ws2.cell(row=i, column=2).font = normal_font
    for c in [1, 2]:
        ws2.cell(row=i, column=c).border = thin_border

ws2.column_dimensions['A'].width = 10
ws2.column_dimensions['B'].width = 25

# Add VLOOKUP hint column in Transaksi sheet for Kategori
ws1.cell(row=1, column=9, value='Kategori')
style_header_row(ws1, 1, 9)
for i in range(3, 7):
    ws1.cell(row=i, column=9, value='')
    ws1.cell(row=i, column=9).border = thin_border
    ws1.cell(row=i, column=9).font = normal_font
ws1.column_dimensions['I'].width = 25

# Hint
ws1.cell(row=summary_row + 8, column=1, value='VLOOKUP di kolom Kategori')
ws1.cell(row=summary_row + 8, column=1).font = Font(name='Calibri', size=9, italic=True, color='888888')

# ============================================================
# SHEET 3: Data Mentah (Soal 5-7)
# ============================================================
ws3 = wb.create_sheet('Data Mentah')

headers3 = ['Tanggal', 'Keterangan', 'Debit', 'Kredit', 'Bulan']
for c, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=c, value=h)
style_header_row(ws3, 1, 5)

# 12 rows of transaction data for pivot
transaksi_banyak = [
    ('03/06/2024', 'Pembelian ATK', 2000000, 0),
    ('05/06/2024', 'Pendapatan Jasa', 0, 5000000),
    ('07/06/2024', 'Bayar Listrik', 500000, 0),
    ('10/06/2024', 'Bayar Gaji', 15000000, 0),
    ('15/06/2024', 'Pendapatan Jasa', 0, 3000000),
    ('20/06/2024', 'Pembelian ATK', 1000000, 0),
    ('25/06/2024', 'Bayar Sewa', 3000000, 0),
    ('02/07/2024', 'Pendapatan Jasa', 0, 7000000),
    ('08/07/2024', 'Bayar Listrik', 600000, 0),
    ('12/07/2024', 'Pembelian ATK', 1500000, 0),
    ('18/07/2024', 'Pendapatan Jasa', 0, 4000000),
    ('22/07/2024', 'Bayar Gaji', 15000000, 0),
]

for i, (tgl, ket, deb, kre) in enumerate(transaksi_banyak, 2):
    ws3.cell(row=i, column=1, value=tgl)
    ws3.cell(row=i, column=1).number_format = date_fmt
    ws3.cell(row=i, column=2, value=ket)
    ws3.cell(row=i, column=3, value=deb)
    ws3.cell(row=i, column=3).number_format = rupiah_fmt
    ws3.cell(row=i, column=4, value=kre)
    ws3.cell(row=i, column=4).number_format = rupiah_fmt
    # Bulan column - blank for candidate to fill with TEXT formula
    ws3.cell(row=i, column=5, value='')
    ws3.cell(row=i, column=5).border = thin_border

style_range(ws3, 2, len(transaksi_banyak)+1, 1, 4)

ws3.column_dimensions['A'].width = 14
ws3.column_dimensions['B'].width = 20
ws3.column_dimensions['C'].width = 16
ws3.column_dimensions['D'].width = 16
ws3.column_dimensions['E'].width = 12

# Hint
hint_row3 = len(transaksi_banyak) + 3
ws3.cell(row=hint_row3, column=1, value='Petunjuk:')
ws3.cell(row=hint_row3, column=1).font = Font(name='Calibri', bold=True, size=9, color='888888')
ws3.cell(row=hint_row3 + 1, column=1, value='Kolom Bulan: =TEXT(A2,"MMMM") — drag ke bawah')
ws3.cell(row=hint_row3 + 1, column=1).font = Font(name='Calibri', size=9, italic=True, color='888888')
ws3.cell(row=hint_row3 + 2, column=1, value='Kemudian buat Pivot Table dari data ini')
ws3.cell(row=hint_row3 + 2, column=1).font = Font(name='Calibri', size=9, italic=True, color='888888')

# ============================================================
# SHEET 4: Buku Perusahaan (Soal 8 - Rekonsiliasi)
# ============================================================
ws4 = wb.create_sheet('Buku Perusahaan')
headers4 = ['No', 'Tanggal', 'Keterangan', 'Nominal']
for c, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=c, value=h)
style_header_row(ws4, 1, 4)

buku_data = [
    (1, '01/06/2024', 'Setoran Tunai', 10000000),
    (2, '03/06/2024', 'Pembelian ATK', 2000000),
    (3, '05/06/2024', 'Pendapatan Jasa', 5000000),
    (4, '07/06/2024', 'Bayar Listrik', 500000),
    (5, '10/06/2024', 'Bayar Gaji', 15000000),
    (6, '12/06/2024', 'Pembelian Komputer', 8000000),
    (7, '15/06/2024', 'Pendapatan Jasa', 3000000),
]

for i, (no, tgl, ket, nominal) in enumerate(buku_data, 2):
    ws4.cell(row=i, column=1, value=no)
    ws4.cell(row=i, column=2, value=tgl)
    ws4.cell(row=i, column=2).number_format = date_fmt
    ws4.cell(row=i, column=3, value=ket)
    ws4.cell(row=i, column=4, value=nominal)
    ws4.cell(row=i, column=4).number_format = rupiah_fmt

style_range(ws4, 2, len(buku_data)+1, 1, 4)

# Kolom Cocok
ws4.cell(row=1, column=5, value='Cocok?')
style_header_row(ws4, 1, 5)
for i in range(2, len(buku_data)+2):
    ws4.cell(row=i, column=5, value='')
    ws4.cell(row=i, column=5).border = thin_border
    ws4.cell(row=i, column=5).font = normal_font

ws4.column_dimensions['A'].width = 6
ws4.column_dimensions['B'].width = 14
ws4.column_dimensions['C'].width = 22
ws4.column_dimensions['D'].width = 16
ws4.column_dimensions['E'].width = 12

hint_row4 = len(buku_data) + 3
ws4.cell(row=hint_row4, column=1, value='Petunjuk:')
ws4.cell(row=hint_row4, column=1).font = Font(name='Calibri', bold=True, size=9, color='888888')
ws4.cell(row=hint_row4 + 1, column=1, value='Rumus: =IF(COUNTIF(Rekening Koran!D:D,D2)>0,"Sesuai","Tidak Sesuai")')
ws4.cell(row=hint_row4 + 1, column=1).font = Font(name='Calibri', size=9, italic=True, color='888888')

# ============================================================
# SHEET 5: Rekening Koran (Soal 8)
# ============================================================
ws5 = wb.create_sheet('Rekening Koran')
headers5 = ['No', 'Tanggal', 'Keterangan', 'Nominal']
for c, h in enumerate(headers5, 1):
    ws5.cell(row=1, column=c, value=h)
style_header_row(ws5, 1, 4)

koran_data = [
    (1, '01/06/2024', 'Setoran Tunai', 10000000),
    (2, '03/06/2024', 'Pembelian ATK', 2000000),
    (3, '05/06/2024', 'Pendapatan Jasa', 5000000),
    (4, '07/06/2024', 'Bayar Listrik', 500000),
    (5, '10/06/2024', 'Bayar Gaji', 15000000),
    (6, '13/06/2024', 'Biaya Admin Bank', 15000),      # hanya di rekening koran
    (7, '14/06/2024', 'Bunga', 25000),                 # hanya di rekening koran
]

for i, (no, tgl, ket, nominal) in enumerate(koran_data, 2):
    ws5.cell(row=i, column=1, value=no)
    ws5.cell(row=i, column=2, value=tgl)
    ws5.cell(row=i, column=2).number_format = date_fmt
    ws5.cell(row=i, column=3, value=ket)
    ws5.cell(row=i, column=4, value=nominal)
    ws5.cell(row=i, column=4).number_format = rupiah_fmt

style_range(ws5, 2, len(koran_data)+1, 1, 4)

ws5.column_dimensions['A'].width = 6
ws5.column_dimensions['B'].width = 14
ws5.column_dimensions['C'].width = 22
ws5.column_dimensions['D'].width = 16

# Highlighted row note
note_row = len(koran_data) + 3
ws5.cell(row=note_row, column=1, value='Catatan:')
ws5.cell(row=note_row, column=1).font = Font(name='Calibri', bold=True, size=9, color='888888')
ws5.cell(row=note_row + 1, column=1, value='Biaya Admin & Bunga hanya ada di Rekening Koran (tidak di Buku Perusahaan)')
ws5.cell(row=note_row + 1, column=1).font = Font(name='Calibri', size=9, italic=True, color='888888')
ws5.cell(row=note_row + 2, column=1, value='Pembelian Komputer hanya ada di Buku Perusahaan (belum direkam bank)')
ws5.cell(row=note_row + 2, column=1).font = Font(name='Calibri', size=9, italic=True, color='888888')

# ============================================================
# SHEET: PETUNJUK (Readme)
# ============================================================
ws6 = wb.create_sheet('Petunjuk')
instructions = [
    'PETUNJUK TEST EXCEL — Finance & Administration',
    '',
    'Durasi: 30-45 menit',
    'Nilai Minimal Lolos: 65/100',
    '',
    '=== SOAL ===',
    '',
    'Soal 1 (15 poin): Data Entry & Format Tabel',
    '  - Isi data di sheet Transaksi sesuai format',
    '  - Buat rumus Saldo: = saldo sebelumnya + (kredit - debit)',
    '',
    'Soal 2 (10 poin): SUM, AVERAGE, COUNT',
    '  - Hitung Total Debit, Total Kredit, Rata-rata Debit, Jumlah Transaksi',
    '  - Letakkan di area RINGKASAN sheet Transaksi',
    '',
    'Soal 3 (15 poin): IF Logic — Status Transaksi',
    '  - Kolom Status: jika Kredit>0 → "Penerimaan", Debit>0 → "Pengeluaran"',
    '',
    'Soal 4 (20 poin): VLOOKUP — Mapping Kategori',
    '  - Gunakan VLOOKUP dari sheet Kategori untuk mengisi kolom Kategori',
    '',
    'Soal 5 (15 poin): Pivot Table / Rekap Bulanan',
    '  - Di sheet Data Mentah, ekstrak Bulan dengan TEXT(A2,"MMMM")',
    '  - Buat Pivot Table: Baris = Bulan, Value = Sum Debit + Sum Kredit',
    '',
    'Soal 6 (10 poin): Conditional Formatting',
    '  - Debit > 5jt → merah; 500rb-5jt → kuning; < 500rb → hijau',
    '',
    'Soal 7 (10 poin): Data Validation',
    '  - Dropdown di kolom Jenis Transaksi: Kas, Bank, Piutang, Utang',
    '',
    'Soal 8 (5 poin): Rekonsiliasi Mini',
    '  - Cocokkan Buku Perusahaan vs Rekening Koran pakai COUNTIF',
    '',
    '=== KETENTUAN ===',
    '',
    '- Simpan file dengan nama: NAMA_CANDIDATE_TestExcel.xlsx',
    '- Jangan hapus atau modifikasi sheet Petunjuk, Kategori, Rekening Koran',
]

for i, line in enumerate(instructions, 1):
    cell = ws6.cell(row=i, column=1, value=line)
    if line.startswith('PETUNJUK'):
        cell.font = Font(name='Calibri', bold=True, size=14, color='2F5496')
    elif line.startswith('Soal'):
        cell.font = Font(name='Calibri', bold=True, size=11)
    elif line.startswith('==='):
        cell.font = Font(name='Calibri', bold=True, size=12, color='2F5496')
    else:
        cell.font = Font(name='Calibri', size=10)

ws6.column_dimensions['A'].width = 100

# ============================================================
# SAVE
# ============================================================
output_path = '/home/ubuntu/meridian/TestExcel_FinanceAdmin_FreshGrad.xlsx'
wb.save(output_path)
print(f'Excel test file created: {output_path}')
print(f'Sheets: {wb.sheetnames}')
