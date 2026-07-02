#!/usr/bin/env python3
"""Generate payroll-focused Excel test for HRD / Finance interview.

Usage: python3 create_excel_test_gaji.py

Requires: openpyxl (pip3 install openpyxl)

Structure:
  Sheet 1: Petunjuk         — candidate instructions
  Sheet 2: Data Karyawan    — employee master data (VLOOKUP source)
  Sheet 3: Daftar Gaji      — main payroll table (formulas to fill)
  Sheet 4: Tarif PPh        — PPh 21 tax bracket lookup table
  Sheet 5: THR              — THR prorata calculation table
  Sheet 6: Rekap Departemen — department summary (SUMIF/COUNTIF)
"""

import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

# ── STYLES ──
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
SUBTLE_FONT = Font(name='Calibri', size=9, italic=True, color='888888')
NORMAL_FONT = Font(name='Calibri', size=11)
BOLD_FONT = Font(name='Calibri', bold=True, size=11)
RUPIAH_FMT = '#,##0'
DATE_FMT = 'DD/MM/YYYY'
PCT_FMT = '0%'
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def style_data(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.font = NORMAL_FONT


def add_hint(ws, row, col, text):
    ws.cell(row=row, column=col, value=text).font = SUBTLE_FONT


# ── SHEET 1: Petunjuk ──
def build_petunjuk(ws):
    lines = [
        'PETUNJUK TEST EXCEL — PENGGAJIAN (PAYROLL)',
        '',
        'Durasi: 30-45 menit | Nilai Minimal Lolos: 65/100',
        '',
        '=== SOAL ===',
        '',
        'Soal 1 (10pt): Data Entry & Format Tabel',
        '  Isi data di sheet Daftar Gaji minimal 8 karyawan.',
        '  Lengkapi: Nama, Departemen, Gaji Pokok, Tunjangan, Status PTKP.',
        '  Format: Rupiah (tanpa desimal), rata tengah, rapi.',
        '',
        'Soal 2 (15pt): Rumus Take Home Pay (THP)',
        '  Kolom H: THP = Gaji Pokok + Tunjangan - BPJS Karyawan - PPh 21',
        '  =F{row}+G{row}-I{row}-J{row}',
        '',
        'Soal 3 (10pt): IF Logic — Status Pajak',
        '  Kolom K: Jika PKP > 0 maka "Kena Pajak", selain itu "Tidak Kena Pajak"',
        '  Gunakan acuan kolom PKP (kolom D di sheet Tarif PPh).',
        '',
        'Soal 4 (15pt): VLOOKUP — Tarif PPh 21',
        '  Kolom J (PPh 21): gunakan VLOOKUP dari sheet Tarif PPh.',
        '  Cari berdasarkan PKP karyawan, ambil tarif, kalikan PKP.',
        '',
        'Soal 5 (15pt): Perhitungan THR Prorata',
        '  Di sheet THR: hitung THR berdasarkan bulan masuk kerja.',
        '  Rumus: (masa_kerja / 12) * gaji_bulanan',
        '',
        'Soal 6 (10pt): Conditional Formatting',
        '  THP < Rp 3.000.000 = merah (isi merah, teks putih)',
        '  THP Rp 3jt - Rp 7jt = kuning (isi kuning, teks hitam)',
        '  THP > Rp 7.000.000 = hijau (isi hijau, teks hitam)',
        '',
        'Soal 7 (10pt): Data Validation Dropdown',
        '  Kolom Departemen: dropdown (Finance, HRD, Marketing, Operasional, IT)',
        '  Kolom Status PTKP: dropdown (TK/0, TK/1, TK/2, TK/3, K/0, K/1, K/2, K/3)',
        '',
        'Soal 8 (15pt): COUNTIF / SUMIF per Departemen',
        '  Di sheet Rekap Departemen: total gaji per dept, jumlah karyawan per dept.',
        '',
        '=== KETENTUAN ===',
        '',
        'Simpan file: NAMA_ANDA_TestGaji.xlsx',
        'Semua sheet jangan dihapus.',
        'Rumus harus bisa di-drag ke bawah (relative reference).',
        '',
        '=== BOBOT NILAI ===',
        '',
        'Data Entry & Format       : 10 pt',
        'Rumus THP                 : 15 pt',
        'IF Logic Status Pajak     : 10 pt',
        'VLOOKUP Tarif PPh 21      : 15 pt',
        'THR Prorata               : 15 pt',
        'Conditional Formatting    : 10 pt',
        'Data Validation Dropdown  : 10 pt',
        'COUNTIF / SUMIF per Dept   : 15 pt',
        '-----------------------------------------',
        'TOTAL                      : 100 pt',
    ]
    for i, line in enumerate(lines, 1):
        cell = ws.cell(row=i, column=1, value=line)
        if line.startswith('PETUNJUK'):
            cell.font = Font(name='Calibri', bold=True, size=14, color='2F5496')
        elif line.startswith('Soal'):
            cell.font = Font(name='Calibri', bold=True, size=11)
        elif line.startswith('==='):
            cell.font = Font(name='Calibri', bold=True, size=12, color='2F5496')
        elif line.startswith('Rumus'):
            cell.font = Font(name='Consolas', size=10, color='333333')
        else:
            cell.font = Font(name='Calibri', size=10)
    ws.column_dimensions['A'].width = 110


# ── SHEET 2: Data Karyawan (master data for VLOOKUP) ──
def build_data_karyawan(ws):
    headers = ['NIK', 'Nama', 'Departemen', 'Gaji Pokok', 'Tunjangan', 'Status PTKP', 'Tanggal Masuk']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    data = [
        ('EMP001', 'Ahmad Fauzi', 'Finance', 8500000, 1500000, 'K/1', '01/02/2022'),
        ('EMP002', 'Siti Nurhaliza', 'HRD', 7000000, 1000000, 'TK/0', '15/06/2023'),
        ('EMP003', 'Budi Santoso', 'Marketing', 6500000, 2000000, 'K/2', '01/09/2021'),
        ('EMP004', 'Dewi Lestari', 'Operasional', 5000000, 500000, 'TK/1', '10/03/2024'),
        ('EMP005', 'Rudi Hermawan', 'IT', 9500000, 2500000, 'K/1', '01/01/2020'),
        ('EMP006', 'Ani Rahmawati', 'HRD', 6000000, 800000, 'TK/0', '20/07/2023'),
        ('EMP007', 'Bayu Pratama', 'Finance', 7500000, 1000000, 'K/0', '01/04/2022'),
        ('EMP008', 'Citra Ayu', 'Marketing', 5500000, 1000000, 'TK/2', '01/11/2023'),
        ('EMP009', 'Doni Kurniawan', 'Operasional', 4500000, 300000, 'TK/0', '05/05/2024'),
        ('EMP010', 'Elsa Fitriani', 'IT', 10000000, 3000000, 'K/3', '01/03/2019'),
        ('EMP011', 'Fajar Nugroho', 'Finance', 8000000, 1200000, 'K/0', '01/08/2022'),
        ('EMP012', 'Gita Permata', 'Marketing', 6000000, 1500000, 'TK/1', '15/01/2024'),
    ]

    for i, (nik, nama, dept, gapok, tunjang, ptkp, tgl) in enumerate(data, 2):
        ws.cell(row=i, column=1, value=nik)
        ws.cell(row=i, column=2, value=nama)
        ws.cell(row=i, column=3, value=dept)
        ws.cell(row=i, column=4, value=gapok).number_format = RUPIAH_FMT
        ws.cell(row=i, column=5, value=tunjang).number_format = RUPIAH_FMT
        ws.cell(row=i, column=6, value=ptkp)
        ws.cell(row=i, column=7, value=tgl).number_format = DATE_FMT

    style_data(ws, 2, 13, 1, len(headers))

    cols = {'A': 12, 'B': 22, 'C': 18, 'D': 16, 'E': 16, 'F': 14, 'G': 18}
    for col_letter, w in cols.items():
        ws.column_dimensions[col_letter].width = w

    add_hint(ws, 15, 1, 'Master data karyawan. Gunakan untuk VLOOKUP.')
    add_hint(ws, 16, 1, 'Kolom NIK bisa dijadikan lookup value.')


# ── SHEET 3: Daftar Gaji (main payroll table — candidate fills formulas) ──
def build_daftar_gaji(ws):
    headers = [
        'No', 'NIK', 'Nama', 'Departemen', 'Gaji Pokok', 'Tunjangan',
        'Status PTKP', 'THP', 'BPJS Karyawan', 'PPh 21', 'Status Pajak'
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    # Pre-fill some data; candidate completes formulas
    data = [
        (1, 'EMP001', 'Ahmad Fauzi', 'Finance', 8500000, 1500000, 'K/1'),
        (2, 'EMP002', 'Siti Nurhaliza', 'HRD', 7000000, 1000000, 'TK/0'),
        (3, 'EMP003', 'Budi Santoso', 'Marketing', 6500000, 2000000, 'K/2'),
        (4, 'EMP004', 'Dewi Lestari', 'Operasional', 5000000, 500000, 'TK/1'),
        (5, 'EMP005', 'Rudi Hermawan', 'IT', 9500000, 2500000, 'K/1'),
        (6, 'EMP006', 'Ani Rahmawati', 'HRD', 6000000, 800000, 'TK/0'),
        (7, 'EMP007', 'Bayu Pratama', 'Finance', 7500000, 1000000, 'K/0'),
        (8, 'EMP008', 'Citra Ayu', 'Marketing', 5500000, 1000000, 'TK/2'),
        (9, 'EMP009', 'Doni Kurniawan', 'Operasional', 4500000, 300000, 'TK/0'),
        (10, 'EMP010', 'Elsa Fitriani', 'IT', 10000000, 3000000, 'K/3'),
        (11, 'EMP011', 'Fajar Nugroho', 'Finance', 8000000, 1200000, 'K/0'),
        (12, 'EMP012', 'Gita Permata', 'Marketing', 6000000, 1500000, 'TK/1'),
    ]

    for i, (no, nik, nama, dept, gapok, tunjang, ptkp) in enumerate(data, 2):
        ws.cell(row=i, column=1, value=no)
        ws.cell(row=i, column=2, value=nik)
        ws.cell(row=i, column=3, value=nama)
        ws.cell(row=i, column=4, value=dept)
        ws.cell(row=i, column=5, value=gapok).number_format = RUPIAH_FMT
        ws.cell(row=i, column=6, value=tunjang).number_format = RUPIAH_FMT
        ws.cell(row=i, column=7, value=ptkp)
        # Col 8: THP — candidate fills formula
        ws.cell(row=i, column=8).number_format = RUPIAH_FMT
        # Col 9: BPJS Karyawan — candidate fills (2% of gapok+tunjangan)
        ws.cell(row=i, column=9).number_format = RUPIAH_FMT
        # Col 10: PPh 21 — candidate fills via VLOOKUP
        ws.cell(row=i, column=10).number_format = RUPIAH_FMT
        # Col 11: Status Pajak — candidate fills IF formula
        ws.cell(row=i, column=11)

    style_data(ws, 2, 13, 1, len(headers))

    # Hint rows
    add_hint(ws, 16, 1, 'Soal 2: THP = Gaji Pokok + Tunjangan - BPJS Karyawan - PPh 21')
    add_hint(ws, 17, 1, 'Soal 3: Status Pajak = IF(PKP>0,"Kena Pajak","Tidak Kena Pajak")')
    add_hint(ws, 18, 1, 'Soal 4: PPh 21 pakai VLOOKUP dari sheet Tarif PPh')
    add_hint(ws, 19, 1, 'BPJS Karyawan = 2% x (Gaji Pokok + Tunjangan)')
    add_hint(ws, 20, 1, 'PKP tahunan = (Gaji Pokok+Tunjangan)*12 - PTKP — lihat sheet Tarif PPh')

    # Data validation for Departemen (col 4)
    dv_dept = DataValidation(
        type='list',
        formula1='"Finance,HRD,Marketing,Operasional,IT"',
        allow_blank=True
    )
    dv_dept.error = 'Pilih dari dropdown'
    dv_dept.prompt = 'Pilih Departemen'
    ws.add_data_validation(dv_dept)
    for i in range(2, 20):
        dv_dept.add(ws.cell(row=i, column=4))

    # Data validation for Status PTKP (col 7)
    dv_ptkp = DataValidation(
        type='list',
        formula1='"TK/0,TK/1,TK/2,TK/3,K/0,K/1,K/2,K/3"',
        allow_blank=True
    )
    dv_ptkp.error = 'Pilih dari dropdown'
    dv_ptkp.prompt = 'Pilih Status PTKP'
    ws.add_data_validation(dv_ptkp)
    for i in range(2, 20):
        dv_ptkp.add(ws.cell(row=i, column=7))

    # Conditional formatting for THP (col 8)
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    red_font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')

    ws.conditional_formatting.add('H2:H20',
        CellIsRule(operator='lessThan', formula=['3000000'], fill=red_fill, font=red_font))
    ws.conditional_formatting.add('H2:H20',
        CellIsRule(operator='between', formula=['3000000', '7000000'], fill=yellow_fill))
    ws.conditional_formatting.add('H2:H20',
        CellIsRule(operator='greaterThan', formula=['7000000'], fill=green_fill))

    cols = {'A': 6, 'B': 12, 'C': 22, 'D': 18, 'E': 16, 'F': 14,
            'G': 14, 'H': 18, 'I': 16, 'J': 16, 'K': 20}
    for col_letter, w in cols.items():
        ws.column_dimensions[col_letter].width = w


# ── SHEET 4: Tarif PPh (tax bracket lookup table) ──
def build_tarif_pph(ws):
    # Section 1: PTKP table
    ws.cell(row=1, column=1, value='TABEL PTKP (Penghasilan Tidak Kena Pajak)')
    ws.cell(row=1, column=1).font = Font(name='Calibri', bold=True, size=12, color='2F5496')

    ptkp_headers = ['Status', 'PTKP Setahun']
    for c, h in enumerate(ptkp_headers, 1):
        ws.cell(row=2, column=c, value=h)
    style_header(ws, 2, 2)

    ptkp_data = [
        ('TK/0', 54000000),
        ('TK/1', 58500000),
        ('TK/2', 63000000),
        ('TK/3', 67500000),
        ('K/0', 58500000),
        ('K/1', 63000000),
        ('K/2', 67500000),
        ('K/3', 72000000),
    ]
    for i, (status, ptkp) in enumerate(ptkp_data, 3):
        ws.cell(row=i, column=1, value=status)
        ws.cell(row=i, column=2, value=ptkp).number_format = RUPIAH_FMT
    style_data(ws, 3, 10, 1, 2)
    add_hint(ws, 12, 1, 'Gunakan VLOOKUP: ambil PTKP berdasarkan Status, lalu hitung PKP=Gaji Setahun-PTKP')

    # Section 2: Tarif Progresif Pasal 17
    ws.cell(row=14, column=1, value='TABEL TARIF PPh 21 (Pasal 17 — Lapisan PKP)')
    ws.cell(row=14, column=1).font = Font(name='Calibri', bold=True, size=12, color='2F5496')

    tarif_headers = ['Lapisan', 'Bawah PKP', 'Atas PKP', 'Tarif']
    for c, h in enumerate(tarif_headers, 1):
        ws.cell(row=15, column=c, value=h)
    style_header(ws, 15, 4)

    tarif_data = [
        ('Lapisan 1', 0, 60000000, 0.05),
        ('Lapisan 2', 60000000, 250000000, 0.15),
        ('Lapisan 3', 250000000, 500000000, 0.25),
        ('Lapisan 4', 500000000, 5000000000, 0.30),
    ]
    for i, (lap, bawah, atas, tarif) in enumerate(tarif_data, 16):
        ws.cell(row=i, column=1, value=lap)
        ws.cell(row=i, column=2, value=bawah).number_format = RUPIAH_FMT
        ws.cell(row=i, column=3, value=atas).number_format = RUPIAH_FMT
        ws.cell(row=i, column=4, value=tarif).number_format = PCT_FMT
    style_data(ws, 16, 19, 1, 4)
    add_hint(ws, 21, 1, 'PKP = (Gaji Pokok+Tunjangan)*12 - PTKP. Lalu hitung PPh setahun pakai tarif progresif.')

    cols = {'A': 14, 'B': 18, 'C': 18, 'D': 10}
    for col_letter, w in cols.items():
        ws.column_dimensions[col_letter].width = w


# ── SHEET 5: THR ──
def build_thr(ws):
    headers = ['No', 'NIK', 'Nama', 'Gaji Bulanan', 'Tanggal Masuk', 'Masa Kerja (bln)', 'THR']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    data = [
        (1, 'EMP004', 'Dewi Lestari', 5500000, '10/03/2024'),
        (2, 'EMP009', 'Doni Kurniawan', 4800000, '05/05/2024'),
        (3, 'EMP012', 'Gita Permata', 7500000, '15/01/2024'),
        (4, 'EMP006', 'Ani Rahmawati', 6800000, '20/07/2023'),
        (5, 'EMP008', 'Citra Ayu', 6500000, '01/11/2023'),
    ]

    for i, (no, nik, nama, gaji, tgl) in enumerate(data, 2):
        ws.cell(row=i, column=1, value=no)
        ws.cell(row=i, column=2, value=nik)
        ws.cell(row=i, column=3, value=nama)
        ws.cell(row=i, column=4, value=gaji).number_format = RUPIAH_FMT
        ws.cell(row=i, column=5, value=tgl).number_format = DATE_FMT
        # Masa Kerja — candidate fills formula (bulan)
        ws.cell(row=i, column=6)
        # THR — candidate fills formula
        ws.cell(row=i, column=7).number_format = RUPIAH_FMT

    style_data(ws, 2, 6, 1, len(headers))

    add_hint(ws, 9, 1, 'Soal 5: Hitung THR Prorata')
    add_hint(ws, 10, 1, 'Masa Kerja = bulan sejak tgl masuk sampai April 2025 (saat THR dibayar)')
    add_hint(ws, 11, 1, 'Rumus: =DATEDIF(E{row},DATE(2025,4,1),"m")')
    add_hint(ws, 12, 1, 'THR = (Masa Kerja/12) x Gaji Bulanan')
    add_hint(ws, 13, 1, 'Jika masa kerja >= 12 bulan, THR = 1 x gaji bulanan')

    cols = {'A': 6, 'B': 12, 'C': 22, 'D': 18, 'E': 18, 'F': 18, 'G': 16}
    for col_letter, w in cols.items():
        ws.column_dimensions[col_letter].width = w


# ── SHEET 6: Rekap Departemen ──
def build_rekap_departemen(ws):
    headers = ['Departemen', 'Total Gaji', 'Rata-rata Gaji', 'Jumlah Karyawan']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    depts = ['Finance', 'HRD', 'Marketing', 'Operasional', 'IT']
    for i, dept in enumerate(depts, 2):
        ws.cell(row=i, column=1, value=dept)
        # Total Gaji — candidate fills SUMIF formula
        ws.cell(row=i, column=2).number_format = RUPIAH_FMT
        # Rata-rata Gaji — candidate fills formula
        ws.cell(row=i, column=3).number_format = RUPIAH_FMT
        # Jumlah Karyawan — candidate fills COUNTIF formula
        ws.cell(row=i, column=4)

    style_data(ws, 2, 6, 1, len(headers))

    add_hint(ws, 8, 1, 'Soal 8: Rekap per Departemen')
    add_hint(ws, 9, 1, 'Total Gaji: =SUMIF(DaftarGaji!D:D, A{row}, DaftarGaji!E:E)')
    add_hint(ws, 10, 1, 'Jumlah Karyawan: =COUNTIF(DaftarGaji!D:D, A{row})')
    add_hint(ws, 11, 1, 'Rata-rata: =B{row}/D{row} atau AVERAGEIF')

    cols = {'A': 20, 'B': 18, 'C': 18, 'D': 20}
    for col_letter, w in cols.items():
        ws.column_dimensions[col_letter].width = w


# ── MAIN ──
def main():
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = 'Petunjuk'
    build_petunjuk(ws1)

    ws2 = wb.create_sheet('Data Karyawan')
    build_data_karyawan(ws2)

    ws3 = wb.create_sheet('Daftar Gaji')
    build_daftar_gaji(ws3)

    ws4 = wb.create_sheet('Tarif PPh')
    build_tarif_pph(ws4)

    ws5 = wb.create_sheet('THR')
    build_thr(ws5)

    ws6 = wb.create_sheet('Rekap Departemen')
    build_rekap_departemen(ws6)

    output = 'TestExcel_Penggajian.xlsx'
    wb.save(output)
    print(f'Done: {output}')
    print(f'Sheets: {wb.sheetnames}')


if __name__ == '__main__':
    main()
