#!/usr/bin/env python3
"""
Generate Excel Test File for Payroll/Salary Interview
TestExcel_Payroll_FreshGrad.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
import os

# --- Styles ---
header_font = Font(bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
hint_font = Font(italic=True, size=10, color='808080')
title_font = Font(bold=True, size=14, color='2F5496')
subtitle_font = Font(bold=True, size=12, color='2F5496')
normal_font = Font(size=11)
bold_font = Font(bold=True, size=11)
rupiah_fmt = '#,##0'
date_fmt = 'DD/MM/YYYY'
pct_fmt = '0%'

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
light_blue_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def style_range(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.font = normal_font
            cell.alignment = Alignment(vertical='center', wrap_text=True)


def set_column_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# =====================================================
# CREATE WORKBOOK
# =====================================================
wb = openpyxl.Workbook()

# =====================================================
# SHEET 1: PETUNJUK
# =====================================================
ws_petunjuk = wb.active
ws_petunjuk.title = "Petunjuk"

ws_petunjuk['A1'] = 'TES EXCEL — PENGGAJIAN (PAYROLL)'
ws_petunjuk['A1'].font = title_font

instructions = [
    '',
    'Waktu: 45 menit',
    '',
    'PETUNJUK UMUM:',
    '1. Kerjakan soal langsung di file Excel ini.',
    '2. Setiap soal sudah disediakan tempatnya di sheet masing-masing.',
    '3. Gunakan rumus Excel (bukan kalkulator manual).',
    '4. Format angka Rupiah tanpa desimal: Rp 1.000.000 (bukan Rp 1.000.000,00)',
    '5. Format tanggal: DD/MM/YYYY',
    '6. Simpan file dengan nama: NamaAnda_TestPayroll.xlsx',
    '',
    'DAFTAR SHEET:',
    '   Data Karyawan   -> Soal 1, 2, 3, 6 (entry data, rumus, IF, conditional format)',
    '   Referensi Gaji  -> Soal 4 (VLOOKUP - tabel referensi)',
    '   Rekap Dept      -> Soal 5, 8 (SUMIF / COUNTIF / pivot)',
    '   Slip Gaji       -> Soal 7 (data validation)',
    '',
    '--- SELAMAT MENGERJAKAN ---',
]

for i, line in enumerate(instructions, 3):
    ws_petunjuk.cell(row=i, column=1, value=line).font = normal_font if i > 3 else bold_font

set_column_widths(ws_petunjuk, [80])

# =====================================================
# SHEET 2: DATA KARYAWAN
# =====================================================
ws_karyawan = wb.create_sheet("Data Karyawan")

# Title
ws_karyawan['A1'] = 'SOAL 1, 2, 3, 6 — Data Karyawan'
ws_karyawan['A1'].font = title_font
ws_karyawan.merge_cells('A1:I1')

ws_karyawan['A2'] = 'Isi kolom yang kosong (kuning) dengan rumus. Gunakan Referensi Gaji untuk VLOOKUP.'
ws_karyawan['A2'].font = hint_font
ws_karyawan.merge_cells('A2:I2')

# Headers (row 4)
headers = ['No', 'Nama Karyawan', 'Departemen', 'Golongan', 'Gaji Pokok', 'Tunjangan Tetap', 'Total Gaji Bruto', 'Status Karyawan', 'BPJS Kesehatan']
for col, h in enumerate(headers, 1):
    ws_karyawan.cell(row=4, column=col, value=h)
style_header_row(ws_karyawan, 4, len(headers))

# Employee data
employees = [
    [1, 'Ani Wijaya',       'Keuangan',   'A', 5500000, '', '', '', ''],
    [2, 'Budi Santoso',     'Marketing',  'B', 4200000, '', '', '', ''],
    [3, 'Citra Dewi',       'HRD',        'A', 6000000, '', '', '', ''],
    [4, 'Deni Pratama',     'IT',         'C', 7800000, '', '', '', ''],
    [5, 'Eka Putri',        'Keuangan',   'B', 4500000, '', '', '', ''],
    [6, 'Fajar Hidayat',    'Marketing',  'A', 5800000, '', '', '', ''],
    [7, 'Gita Permata',     'HRD',        'C', 7200000, '', '', '', ''],
    [8, 'Hendra Gunawan',   'IT',         'B', 4800000, '', '', '', ''],
    [9, 'Indah Lestari',    'Keuangan',   'A', 5300000, '', '', '', ''],
    [10,'Joko Widodo',      'Marketing',  'C', 8500000, '', '', '', ''],
]

for i, emp in enumerate(employees):
    row = 5 + i
    for col, val in enumerate(emp, 1):
        cell = ws_karyawan.cell(row=row, column=col, value=val)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')
        if col == 1:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        if col in (5, 6, 7, 9):  # currency columns
            cell.number_format = rupiah_fmt
    # Color the formula columns (F, G, H, I) - light yellow
    for col in (6, 7, 8, 9):
        ws_karyawan.cell(row=row, column=col).fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

style_range(ws_karyawan, 5, 14, 1, 9)
set_column_widths(ws_karyawan, [5, 18, 13, 10, 15, 15, 18, 18, 15])

# Formula hints
ws_karyawan['B16'] = 'Rumus yang harus diisi:'
ws_karyawan['B16'].font = bold_font
hints = [
    'F5:F14 = VLOOKUP(Golongan, Referensi Gaji!A:B, 2, FALSE) — mencari tunjangan tetap',
    'G5:G14 = Gaji Pokok + Tunjangan Tetap (E + F)',
    'H5:H14 = IF(Gaji Pokok > 5.000.000, "Senior", "Junior")',
    'I5:I14 = Gaji Pokok * 1% (BPJS Kesehatan tarif 1% dari gaji pokok)',
]
for i, hint in enumerate(hints):
    ws_karyawan.cell(row=17 + i, column=2, value=hint).font = hint_font

# Soal 2 area
ws_karyawan['A20'] = 'SOAL 2 — SUM, AVERAGE, COUNT'
ws_karyawan['A20'].font = subtitle_font
ws_karyawan.merge_cells('A20:I20')

stats_headers = ['Statistik', 'Nilai', 'Rumus']
for col, h in enumerate(stats_headers, 1):
    ws_karyawan.cell(row=21, column=col, value=h)
style_header_row(ws_karyawan, 21, 3)

stats_data = [
    ['Total Beban Gaji Bruto', '', '=SUM(G5:G14)'],
    ['Rata-rata Gaji Karyawan', '', '=AVERAGE(G5:G14)'],
    ['Jumlah Karyawan',         '', '=COUNT(A5:A14)'],
    ['Karyawan Senior',         '', '=COUNTIF(H5:H14,"Senior")'],
    ['Karyawan Junior',         '', '=COUNTIF(H5:H14,"Junior")'],
]
for i, stat in enumerate(stats_data):
    row = 22 + i
    for col, val in enumerate(stat, 1):
        cell = ws_karyawan.cell(row=row, column=col, value=val)
        cell.font = normal_font
        cell.border = thin_border
        if col == 2:
            cell.number_format = rupiah_fmt
            cell.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

# Soal 6 area - conditional formatting
ws_karyawan['A28'] = 'SOAL 6 — Conditional Formatting (diterapkan di range G5:G14)'
ws_karyawan['A28'].font = subtitle_font
ws_karyawan.merge_cells('A28:I28')
ws_karyawan['A29'] = 'Beri format: Hijau untuk nilai tertinggi, Merah untuk terendah, Kuning untuk di atas rata-rata'
ws_karyawan['A29'].font = hint_font
ws_karyawan.merge_cells('A29:I29')

# Apply conditional formatting rules
ws_karyawan.conditional_formatting.add(
    'G5:G14',
    CellIsRule(operator='greaterThan', formula=['AVERAGE($G$5:$G$14)'], fill=yellow_fill)
)

# =====================================================
# SHEET 3: REFERENSI GAJI
# =====================================================
ws_ref = wb.create_sheet("Referensi Gaji")

ws_ref['A1'] = 'SOAL 4 — Tabel Referensi Tunjangan (untuk VLOOKUP)'
ws_ref['A1'].font = title_font
ws_ref.merge_cells('A1:C1')
ws_ref['A2'] = 'Gunakan tabel ini untuk mencari tunjangan tetap berdasarkan golongan'
ws_ref['A2'].font = hint_font
ws_ref.merge_cells('A2:C2')

ref_headers = ['Golongan', 'Tunjangan Tetap', 'Keterangan']
for col, h in enumerate(ref_headers, 1):
    ws_ref.cell(row=4, column=col, value=h)
style_header_row(ws_ref, 4, 3)

ref_data = [
    ['A', 1500000, 'Senior Staff'],
    ['B', 1000000, 'Staff'],
    ['C', 2000000, 'Supervisor / Team Lead'],
    ['D', 3000000, 'Manager'],
]

for i, ref in enumerate(ref_data):
    row = 5 + i
    for col, val in enumerate(ref, 1):
        cell = ws_ref.cell(row=row, column=col, value=val)
        cell.font = normal_font
        cell.border = thin_border
        if col == 2:
            cell.number_format = rupiah_fmt

set_column_widths(ws_ref, [12, 18, 25])

# =====================================================
# SHEET 4: REKAP DEPT
# =====================================================
ws_rekap = wb.create_sheet("Rekap Dept")

ws_rekap['A1'] = 'SOAL 5 & 8 — Rekap Gaji per Departemen'
ws_rekap['A1'].font = title_font
ws_rekap.merge_cells('A1:E1')

ws_rekap['A2'] = 'Buat rekap menggunakan rumus SUMIF / COUNTIF berdasarkan data di sheet Data Karyawan'
ws_rekap['A2'].font = hint_font
ws_rekap.merge_cells('A2:E2')

# Part A: SUMIF
ws_rekap['A4'] = 'SOAL 5a — Total Gaji per Departemen (SUMIF)'
ws_rekap['A4'].font = subtitle_font
ws_rekap.merge_cells('A4:C4')

rekap_headers = ['Departemen', 'Total Gaji Bruto', 'Rumus yang digunakan']
for col, h in enumerate(rekap_headers, 1):
    ws_rekap.cell(row=5, column=col, value=h)
style_header_row(ws_rekap, 5, 3)

departments = ['Keuangan', 'Marketing', 'HRD', 'IT']
for i, dept in enumerate(departments):
    row = 6 + i
    ws_rekap.cell(row=row, column=1, value=dept).font = normal_font
    ws_rekap.cell(row=row, column=1).border = thin_border
    cell_val = ws_rekap.cell(row=row, column=2)
    cell_val.value = ''  # Candidate fills
    cell_val.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    cell_val.number_format = rupiah_fmt
    cell_val.border = thin_border
    hint_cell = ws_rekap.cell(row=row, column=3, value='=SUMIF(Data Karyawan!C:C, A{0}, Data Karyawan!G:G)'.format(row))
    hint_cell.font = hint_font
    hint_cell.border = thin_border

# Part B: COUNTIF
ws_rekap['A11'] = 'SOAL 8 — Jumlah Karyawan per Departemen (COUNTIF)'
ws_rekap['A11'].font = subtitle_font
ws_rekap.merge_cells('A11:C11')

count_headers = ['Departemen', 'Jumlah Karyawan', 'Rumus yang digunakan']
for col, h in enumerate(count_headers, 1):
    ws_rekap.cell(row=12, column=col, value=h)
style_header_row(ws_rekap, 12, 3)

for i, dept in enumerate(departments):
    row = 13 + i
    ws_rekap.cell(row=row, column=1, value=dept).font = normal_font
    ws_rekap.cell(row=row, column=1).border = thin_border
    cell_val = ws_rekap.cell(row=row, column=2)
    cell_val.value = ''  # Candidate fills
    cell_val.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    cell_val.border = thin_border
    hint_cell = ws_rekap.cell(row=row, column=3, value='=COUNTIF(Data Karyawan!C:C, A{0})'.format(row))
    hint_cell.font = hint_font
    hint_cell.border = thin_border

set_column_widths(ws_rekap, [15, 20, 50])

# =====================================================
# SHEET 5: SLIP GAJI
# =====================================================
ws_slip = wb.create_sheet("Slip Gaji")

ws_slip['A1'] = 'SOAL 7 — Slip Gaji Karyawan (Data Validation)'
ws_slip['A1'].font = title_font
ws_slip.merge_cells('A1:G1')

ws_slip['A2'] = 'Buat slip gaji untuk 1 karyawan. Gunakan dropdown validasi untuk kolom Departemen & Golongan.'
ws_slip['A2'].font = hint_font
ws_slip.merge_cells('A2:G2')

# Slip header
ws_slip['A4'] = 'SLIP GAJI — BULAN INI'
ws_slip['A4'].font = subtitle_font
ws_slip.merge_cells('A4:C4')

# Slip structure
slip_items = [
    ['Nama Karyawan', 'Ani Wijaya'],
    ['Departemen', ''],  # Dropdown
    ['Golongan', ''],    # Dropdown
    ['Gaji Pokok', 5500000],
    ['Tunjangan Tetap', ''],  # VLOOKUP
    ['Total Gaji Bruto', ''], # Formula
    [''],
    ['POTONGAN:'],
    ['BPJS Kesehatan (1%)', ''],  # 1% of gaji pokok
    ['BPJS Ketenagakerjaan (JHT 2%)', ''],  # 2% of gaji pokok
    ['Total Potongan', ''],
    [''],
    ['GAJI BERSIH (NETO)', ''],  # Bruto - Potongan
]

for i, item in enumerate(slip_items):
    row = 6 + i
    if len(item) == 0:
        continue
    if len(item) == 1:
        ws_slip.cell(row=row, column=1, value=item[0]).font = bold_font
        continue
    ws_slip.cell(row=row, column=1, value=item[0]).font = normal_font
    ws_slip.cell(row=row, column=1).border = thin_border
    
    cell = ws_slip.cell(row=row, column=2, value=item[1])
    cell.font = normal_font
    cell.border = thin_border
    if isinstance(item[1], (int, float)) and item[1] > 1000:
        cell.number_format = rupiah_fmt
    
    # Mark empty cells yellow
    if item[1] == '' or (isinstance(item[1], str) and item[1] == ''):
        cell.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

# Add data validation for Departemen (B7)
dv_dept = DataValidation(type="list", formula1='"Keuangan,Marketing,HRD,IT,Operasional"', allow_blank=True)
dv_dept.error = "Pilih departemen yang tersedia"
dv_dept.errorTitle = "Departemen tidak valid"
ws_slip.add_data_validation(dv_dept)
dv_dept.add('B7')

# Add data validation for Golongan (B8)
dv_gol = DataValidation(type="list", formula1='"A,B,C,D"', allow_blank=True)
dv_gol.error = "Pilih golongan yang tersedia"
dv_gol.errorTitle = "Golongan tidak valid"
ws_slip.add_data_validation(dv_gol)
dv_gol.add('B8')

# Hints
ws_slip['E6'] = 'Petunjuk:'
ws_slip['E6'].font = bold_font
hints_slip = [
    'B7: gunakan dropdown validasi',
    'B8: gunakan dropdown validasi',
    'B10: =VLOOKUP(B8, Referensi Gaji!A:B, 2, FALSE)',
    'B11: =B9+B10',
    'B14: =B9*1% (BPJS Kesehatan)',
    'B15: =B9*2% (BPJS JHT Karyawan)',
    'B16: =B14+B15',
    'B18: =B11-B16  (Gaji Bersih)',
]
for i, h in enumerate(hints_slip):
    ws_slip.cell(row=7 + i, column=5, value=h).font = hint_font

set_column_widths(ws_slip, [22, 20, 5, 3, 45])

# =====================================================
# SAVE
# =====================================================
output_path = '/home/ubuntu/meridian/TestExcel_Payroll_FreshGrad.xlsx'
wb.save(output_path)
print(f"Excel file saved to: {output_path}")
print(f"File size: {os.path.getsize(output_path):,} bytes")
