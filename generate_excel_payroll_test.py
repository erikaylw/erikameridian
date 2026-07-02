#!/usr/bin/env python3
"""Generate Excel test for Payroll recruitment — HRD / Finance position.

Usage: python3 generate_excel_payroll_test.py

Sheets:
  - Transaksi Gaji    — soal 1, 3, 6, 7
  - Level Gaji        — VLOOKUP lookup table, soal 4
  - Data Karyawan     — employee master + VLOOKUP exercise, soal 4, 6
  - Data Mentah Gaji  — raw payroll data for pivot, soal 5
  - Rekonsiliasi Gaji — COUNTIF reconciliation, soal 8
  - Petunjuk          — candidate instructions
"""

import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

# ── STYLES ──
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
SUBTLE_FONT = Font(name='Calibri', size=9, italic=True, color='888888')
NORMAL_FONT = Font(name='Calibri', size=11)
BOLD_FONT = Font(name='Calibri', bold=True, size=11)
TITLE_FONT = Font(name='Calibri', bold=True, size=14, color='2F5496')
SECTION_FONT = Font(name='Calibri', bold=True, size=12, color='2F5496')

RUPIAH_FMT = '#,##0'
DATE_FMT = 'DD/MM/YYYY'
PERCENT_FMT = '0%'

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def style_data(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).font = NORMAL_FONT


def add_hint(ws, row, col, text):
    ws.cell(row=row, column=col, value=text).font = SUBTLE_FONT


# ── SHEET 1: Transaksi Gaji ─────────────────────────────────
def build_transaksi(ws):
    headers = ['Tanggal', 'Keterangan', 'Penerimaan', 'Pengeluaran', 'Saldo',
               'Jenis Pembayaran', 'Jenis']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    data = [
        ('01/03/2025', 'Saldo Awal Kas Gaji', 50000000, None),
        ('03/03/2025', 'Transfer Gaji Karyawan', None, 133000000),
        ('05/03/2025', 'Setor PPh 21', None, 12500000),
        ('06/03/2025', 'Setor BPJS', None, 4500000),
        ('10/03/2025', 'Transfer Tunjangan Lembur', None, 5000000),
        ('15/03/2025', 'Top Up Kas Gaji', 150000000, None),
        ('20/03/2025', 'Bayar THR', None, 25000000),
        ('25/03/2025', 'Setor PPh 21 THR', None, 2500000),
        ('28/03/2025', 'Klaim BPJS Reimburse', 3000000, None),
        ('31/03/2025', 'Saldo Akhir', None, None),
    ]

    for i, (tgl, ket, penerimaan, pengeluaran) in enumerate(data, 2):
        ws.cell(row=i, column=1, value=tgl).number_format = DATE_FMT
        ws.cell(row=i, column=2, value=ket)
        if penerimaan is not None:
            ws.cell(row=i, column=3, value=penerimaan).number_format = RUPIAH_FMT
        if pengeluaran is not None:
            ws.cell(row=i, column=4, value=pengeluaran).number_format = RUPIAH_FMT
        # Saldo (col 5) = placeholder
        ws.cell(row=i, column=5).number_format = RUPIAH_FMT
        ws.cell(row=i, column=5).border = THIN_BORDER
        # Jenis Pembayaran (col 6) — candidate fills
        ws.cell(row=i, column=6).border = THIN_BORDER
        # Jenis (col 7) — IF formula placeholder
        ws.cell(row=i, column=7).border = THIN_BORDER

    style_data(ws, 2, 11, 1, 5)

    # Set saldo row 2 = penerimaan (saldo awal)
    ws.cell(row=2, column=5, value=50000000).number_format = RUPIAH_FMT

    # Hints
    add_hint(ws, 2, 8, 'Isi: 50.000.000 (saldo awal)')
    add_hint(ws, 3, 8, 'Rumus Saldo: =E2+(C3-D3) — drag ke bawah')
    add_hint(ws, 12, 1, 'Soal 3 — Kolom Jenis: =IF(C{n}>0,"Pemasukan",IF(D{n}>0,"Pengeluaran",""))')

    # Data Validation for Jenis Pembayaran (col 6)
    dv = DataValidation(type='list', formula1='"Transfer Bank,Tunai,Cek/BG"', allow_blank=True)
    dv.error = 'Pilih dari dropdown'
    dv.prompt = 'Pilih jenis pembayaran'
    dv.promptTitle = 'Jenis Pembayaran'
    ws.add_data_validation(dv)
    for i in range(2, 14):
        dv.add(ws.cell(row=i, column=6))

    # Conditional formatting for Pengeluaran (col 4) — Soal 6
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    red_font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')

    ws.conditional_formatting.add('D2:D13',
        CellIsRule(operator='greaterThan', formula=['50000000'], fill=red_fill, font=red_font))
    ws.conditional_formatting.add('D2:D13',
        CellIsRule(operator='between', formula=['5000000', '50000000'], fill=yellow_fill))
    ws.conditional_formatting.add('D2:D13',
        CellIsRule(operator='lessThan', formula=['5000000'], fill=green_fill))

    # Summary area
    row_start = 14
    ws.cell(row=row_start, column=1, value='RINGKASAN (Soal 2)').font = Font(
        name='Calibri', bold=True, size=12, color='2F5496')
    for j, label in enumerate(['Total Penerimaan', 'Total Pengeluaran',
                                'Rata-rata Pengeluaran', 'Jumlah Transaksi Pengeluaran'], 1):
        ws.cell(row=row_start + j, column=1, value=label).font = BOLD_FONT
        ws.cell(row=row_start + j, column=2).border = THIN_BORDER
        ws.cell(row=row_start + j, column=2).number_format = RUPIAH_FMT

    add_hint(ws, row_start + 1, 4, '=SUM(C2:C11)')
    add_hint(ws, row_start + 2, 4, '=SUM(D2:D11)')
    add_hint(ws, row_start + 3, 4, '=AVERAGE(D2:D11)')
    add_hint(ws, row_start + 4, 4, '=COUNT(D2:D11)')

    # Column widths
    for col_letter, w in {'A': 14, 'B': 28, 'C': 18, 'D': 18, 'E': 18,
                          'F': 20, 'G': 16, 'H': 42}.items():
        ws.column_dimensions[col_letter].width = w


# ── SHEET 2: Level Gaji ────────────────────────────────────
def build_level_gaji(ws):
    headers = ['Level', 'Gaji Pokok', 'Tunjangan']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, 3)

    data = [
        ('Staff', 5000000, 500000),
        ('Senior', 7500000, 750000),
        ('Supervisor', 10000000, 1000000),
        ('Manager', 15000000, 1500000),
    ]
    for i, (level, gapok, tunj) in enumerate(data, 2):
        ws.cell(row=i, column=1, value=level)
        ws.cell(row=i, column=2, value=gapok).number_format = RUPIAH_FMT
        ws.cell(row=i, column=3, value=tunj).number_format = RUPIAH_FMT
    style_data(ws, 2, 5, 1, 3)

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16


# ── SHEET 3: Data Karyawan ──────────────────────────────────
def build_data_karyawan(ws):
    headers = ['NIK', 'Nama', 'Level', 'Gaji Pokok', 'Tunjangan', 'Gaji Bruto',
               'Departemen', 'Status']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    data = [
        ('001', 'Andi', 'Staff', None, None, None, 'Finance', 'Aktif'),
        ('002', 'Dewi', 'Senior', None, None, None, 'Marketing', 'Aktif'),
        ('003', 'Budi', 'Staff', None, None, None, 'Finance', 'Aktif'),
        ('004', 'Sari', 'Supervisor', None, None, None, 'HRD', 'Aktif'),
        ('005', 'Rudi', 'Manager', None, None, None, 'Finance', 'Aktif'),
        ('006', 'Lina', 'Staff', None, None, None, 'Marketing', 'Aktif'),
        ('007', 'Toni', 'Senior', None, None, None, 'HRD', 'Aktif'),
        ('008', 'Maya', 'Supervisor', None, None, None, 'Finance', 'Aktif'),
    ]
    for i, (nik, nama, level, gapok, tunj, brut, dept, status) in enumerate(data, 2):
        ws.cell(row=i, column=1, value=nik).border = THIN_BORDER
        ws.cell(row=i, column=2, value=nama).border = THIN_BORDER
        ws.cell(row=i, column=3, value=level).border = THIN_BORDER
        # Gaji Pokok (col 4) — VLOOKUP placeholder
        ws.cell(row=i, column=4).number_format = RUPIAH_FMT
        ws.cell(row=i, column=4).border = THIN_BORDER
        # Tunjangan (col 5) — VLOOKUP placeholder
        ws.cell(row=i, column=5).number_format = RUPIAH_FMT
        ws.cell(row=i, column=5).border = THIN_BORDER
        # Gaji Bruto (col 6) — formula placeholder
        ws.cell(row=i, column=6).number_format = RUPIAH_FMT
        ws.cell(row=i, column=6).border = THIN_BORDER
        ws.cell(row=i, column=7, value=dept).border = THIN_BORDER
        ws.cell(row=i, column=8, value=status).border = THIN_BORDER
    style_data(ws, 2, 9, 1, 8)

    # Ensure data area has borders (some cells already set above)
    n = Font(name='Calibri', size=11)
    for r in range(2, 10):
        for c in range(1, 9):
            ws.cell(row=r, column=c).font = n

    add_hint(ws, 2, 9, 'Rumus G.Pokok: =VLOOKUP(C2,Level Gaji!$A$2:$C$5,2,FALSE)')
    add_hint(ws, 2, 10, 'Rumus Tunjangan: =VLOOKUP(C2,Level Gaji!$A$2:$C$5,3,FALSE)')
    add_hint(ws, 2, 11, 'Rumus G.Bruto: =D2+E2')

    # Conditional formatting on Gaji Bruto (col 6) — Soal 6
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    red_font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')

    ws.conditional_formatting.add('F2:F9',
        CellIsRule(operator='greaterThan', formula=['15000000'], fill=red_fill, font=red_font))
    ws.conditional_formatting.add('F2:F9',
        CellIsRule(operator='between', formula=['8000000', '15000000'], fill=yellow_fill))
    ws.conditional_formatting.add('F2:F9',
        CellIsRule(operator='lessThan', formula=['8000000'], fill=green_fill))

    for col_letter, w in {'A': 10, 'B': 16, 'C': 16, 'D': 16,
                          'E': 16, 'F': 16, 'G': 16, 'H': 12, 'I': 50, 'J': 30, 'K': 20}.items():
        ws.column_dimensions[col_letter].width = w


# ── SHEET 4: Data Mentah Gaji ───────────────────────────────
def build_data_mentah(ws):
    headers = ['Tanggal', 'NIK', 'Nama', 'Departemen', 'Gaji Bruto',
               'Potongan PPh 21', 'Potongan BPJS', 'Gaji Bersih', 'Bulan']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    data = [
        ('03/03/2025', '001', 'Andi', 'Finance', 5500000, 50000, 200000, 5250000),
        ('03/03/2025', '002', 'Dewi', 'Marketing', 8250000, 75000, 315000, 7860000),
        ('03/03/2025', '003', 'Budi', 'Finance', 5500000, 50000, 200000, 5250000),
        ('03/03/2025', '004', 'Sari', 'HRD', 11000000, 500000, 420000, 10080000),
        ('03/03/2025', '005', 'Rudi', 'Finance', 16500000, 1250000, 630000, 14620000),
        ('03/03/2025', '006', 'Lina', 'Marketing', 5500000, 50000, 200000, 5250000),
        ('03/03/2025', '007', 'Toni', 'HRD', 8250000, 75000, 315000, 7860000),
        ('03/03/2025', '008', 'Maya', 'Finance', 11000000, 500000, 420000, 10080000),
        ('20/03/2025', '001', 'Andi', 'Finance', 2000000, 0, 0, 2000000),
        ('20/03/2025', '002', 'Dewi', 'Marketing', 2000000, 0, 0, 2000000),
        ('20/03/2025', '003', 'Budi', 'Finance', 2000000, 0, 0, 2000000),
        ('20/03/2025', '004', 'Sari', 'HRD', 5000000, 0, 0, 5000000),
        ('20/03/2025', '005', 'Rudi', 'Finance', 7500000, 0, 0, 7500000),
        ('20/03/2025', '006', 'Lina', 'Marketing', 2000000, 0, 0, 2000000),
        ('20/03/2025', '007', 'Toni', 'HRD', 2000000, 0, 0, 2000000),
        ('20/03/2025', '008', 'Maya', 'Finance', 5000000, 0, 0, 5000000),
    ]
    for i, (tgl, nik, nama, dept, brut, ppn21, bpjs, bersih) in enumerate(data, 2):
        ws.cell(row=i, column=1, value=tgl).number_format = DATE_FMT
        ws.cell(row=i, column=2, value=nik)
        ws.cell(row=i, column=3, value=nama)
        ws.cell(row=i, column=4, value=dept)
        ws.cell(row=i, column=5, value=brut).number_format = RUPIAH_FMT
        ws.cell(row=i, column=6, value=ppn21).number_format = RUPIAH_FMT
        ws.cell(row=i, column=7, value=bpjs).number_format = RUPIAH_FMT
        ws.cell(row=i, column=8, value=bersih).number_format = RUPIAH_FMT
        # Bulan (col 9) — leave for candidate
        ws.cell(row=i, column=9).border = THIN_BORDER

    style_data(ws, 2, 17, 1, 9)

    add_hint(ws, 19, 1, 'Kolom Bulan: =TEXT(A2,"MMMM")')
    add_hint(ws, 20, 1, 'Buat Pivot Table: Baris=Departemen, Kolom=Bulan, Value=Sum Gaji Bruto')

    for col_letter, w in {'A': 14, 'B': 8, 'C': 12, 'D': 16, 'E': 16,
                          'F': 18, 'G': 16, 'H': 16, 'I': 14}.items():
        ws.column_dimensions[col_letter].width = w


# ── SHEET 5: Rekonsiliasi Gaji ─────────────────────────────
def build_rekonsiliasi(ws):
    # Section 1: Daftar Karyawan
    ws.cell(row=1, column=1, value='DATA KARYAWAN').font = SECTION_FONT
    headers1 = ['No', 'NIK', 'Nama', 'Gaji Bruto', 'Status Bayar']
    for c, h in enumerate(headers1, 1):
        ws.cell(row=2, column=c, value=h)
    style_header(ws, 2, 5)

    data_karyawan = [
        (1, '001', 'Andi', 5500000, ''),
        (2, '002', 'Dewi', 8250000, ''),
        (3, '003', 'Budi', 5500000, ''),
        (4, '004', 'Sari', 11000000, ''),
        (5, '005', 'Rudi', 16500000, ''),
        (6, '006', 'Lina', 5500000, ''),
        (7, '007', 'Toni', 8250000, ''),
        (8, '008', 'Maya', 11000000, ''),
    ]
    for i, (no, nik, nama, brut, status) in enumerate(data_karyawan, 3):
        ws.cell(row=i, column=1, value=no)
        ws.cell(row=i, column=2, value=nik)
        ws.cell(row=i, column=3, value=nama)
        ws.cell(row=i, column=4, value=brut).number_format = RUPIAH_FMT
        ws.cell(row=i, column=5).border = THIN_BORDER  # Status Bayar formula

    style_data(ws, 3, 10, 1, 4)

    # Section 2: Data Pembayaran (yang sudah dibayar)
    start_pembayaran = 13
    ws.cell(row=start_pembayaran, column=1, value='DATA PEMBAYARAN').font = SECTION_FONT
    headers2 = ['No', 'NIK', 'Nama', 'Jumlah Dibayar', 'Tanggal Bayar']
    for c, h in enumerate(headers2, 1):
        ws.cell(row=start_pembayaran + 1, column=c, value=h)
    style_header(ws, start_pembayaran + 1, 5)

    pembayaran = [
        (1, '001', 'Andi', 5250000, '03/03/2025'),
        (2, '002', 'Dewi', 7860000, '03/03/2025'),
        (3, '003', 'Budi', 5250000, '03/03/2025'),
        (4, '004', 'Sari', 10080000, '03/03/2025'),
        (5, '006', 'Lina', 5250000, '03/03/2025'),
        (6, '007', 'Toni', 7860000, '03/03/2025'),
    ]
    for i, (no, nik, nama, jml, tgl) in enumerate(pembayaran, start_pembayaran + 2):
        ws.cell(row=i, column=1, value=no)
        ws.cell(row=i, column=2, value=nik)
        ws.cell(row=i, column=3, value=nama)
        ws.cell(row=i, column=4, value=jml).number_format = RUPIAH_FMT
        ws.cell(row=i, column=5, value=tgl).number_format = DATE_FMT
    style_data(ws, start_pembayaran + 2, start_pembayaran + 7, 1, 5)

    # Section 3: Rekapan COUNTIF
    start_rekap = 22
    ws.cell(row=start_rekap, column=1, value='REKAP STATUS PEMBAYARAN').font = SECTION_FONT
    for c, h in enumerate(['Status', 'Jumlah Karyawan'], 1):
        ws.cell(row=start_rekap + 1, column=c, value=h)
    style_header(ws, start_rekap + 1, 2)

    ws.cell(row=start_rekap + 2, column=1, value='Sudah Dibayar')
    ws.cell(row=start_rekap + 2, column=2).border = THIN_BORDER
    ws.cell(row=start_rekap + 3, column=1, value='Belum Dibayar')
    ws.cell(row=start_rekap + 3, column=2).border = THIN_BORDER
    style_data(ws, start_rekap + 2, start_rekap + 3, 1, 2)

    add_hint(ws, start_rekap + 2, 4, '=COUNTIF(E3:E10,"Sudah")')
    add_hint(ws, start_rekap + 3, 4, '=COUNTIF(E3:E10,"Belum")')
    add_hint(ws, 11, 1, 'Rumus Status: =IF(COUNTIF(B15:B20,B3)>0,"Sudah","Belum")')

    for col_letter, w in {'A': 8, 'B': 10, 'C': 16, 'D': 20, 'E': 16, 'F': 16}.items():
        ws.column_dimensions[col_letter].width = w


# ── SHEET 6: Petunjuk ──────────────────────────────────────
def build_petunjuk(ws):
    lines = [
        'PETUNJUK TEST EXCEL — PENGGAJIAN (PAYROLL)',
        '',
        'Durasi: 30-45 menit | Nilai Minimal Lolos: 65/100',
        '',
        '=== SOAL ===',
        '',
        'Soal 1 (15pt) — Data Entry & Rumus Saldo (sheet Transaksi Gaji)',
        '  a) Isi saldo awal baris 2 = 50.000.000',
        '  b) Rumus saldo: =E{sebelum}+(C{ini}-D{ini}) — drag ke bawah',
        '',
        'Soal 2 (10pt) — SUM, AVERAGE, COUNT (sheet Transaksi Gaji, area Ringkasan)',
        '  Total Penerimaan, Total Pengeluaran, Rata-rata Pengeluaran, Jumlah Transaksi',
        '',
        'Soal 3 (15pt) — IF Logic (sheet Transaksi Gaji, kolom Jenis)',
        '  Rumus: =IF(C{n}>0,"Pemasukan",IF(D{n}>0,"Pengeluaran",""))',
        '',
        'Soal 4 (20pt) — VLOOKUP (sheet Data Karyawan & Level Gaji)',
        '  a) Gaji Pokok = VLOOKUP dari sheet Level Gaji',
        '  b) Tunjangan = VLOOKUP kolom ke-3',
        '  c) Gaji Bruto = Gaji Pokok + Tunjangan',
        '',
        'Soal 5 (15pt) — Pivot Table (sheet Data Mentah Gaji)',
        '  a) Ekstrak bulan dengan =TEXT(A{n},"MMMM")',
        '  b) Buat Pivot Table: Baris=Departemen, Value=Sum Gaji Bruto',
        '',
        'Soal 6 (10pt) — Conditional Formatting (sheet Data Karyawan)',
        '  Gaji Bruto > 15jt = merah, 8jt-15jt = kuning, < 8jt = hijau',
        '',
        'Soal 7 (10pt) — Data Validation Dropdown (sheet Transaksi Gaji)',
        '  Kolom Jenis Pembayaran: Transfer Bank, Tunai, Cek/BG',
        '',
        'Soal 8 (5pt) — COUNTIF Rekonsiliasi (sheet Rekonsiliasi Gaji)',
        '  a) Kolom Status: Sudah atau Belum dibayar',
        '  b) COUNTIF rekap jumlah Sudah dan Belum',
        '',
        '=== KETENTUAN ===',
        '',
        'Simpan: NAMA_CANDIDATE_TestExcel_Payroll.xlsx',
        'Jangan edit atau hapus sheet: Level Gaji, Petunjuk',
        'Format Rupiah tanpa desimal untuk semua nominal',
        'Gunakan format tanggal DD/MM/YYYY',
        '',
        'Selamat mengerjakan!',
    ]
    for i, line in enumerate(lines, 1):
        cell = ws.cell(row=i, column=1, value=line)
        if line.startswith('PETUNJUK'):
            cell.font = TITLE_FONT
        elif line.startswith('Soal'):
            cell.font = Font(name='Calibri', bold=True, size=11)
        elif line.startswith('==='):
            cell.font = SECTION_FONT
        elif line.startswith('  a)') or line.startswith('  b)') or line.startswith('  c)'):
            cell.font = Font(name='Calibri', size=10, italic=True, color='444444')
        else:
            cell.font = Font(name='Calibri', size=10)
    ws.column_dimensions['A'].width = 110


# ── MAIN ────────────────────────────────────────────────────
def main():
    wb = openpyxl.Workbook()

    wb.active.title = 'Transaksi Gaji'
    build_transaksi(wb['Transaksi Gaji'])

    ws2 = wb.create_sheet('Level Gaji')
    build_level_gaji(ws2)

    ws3 = wb.create_sheet('Data Karyawan')
    build_data_karyawan(ws3)

    ws4 = wb.create_sheet('Data Mentah Gaji')
    build_data_mentah(ws4)

    ws5 = wb.create_sheet('Rekonsiliasi Gaji')
    build_rekonsiliasi(ws5)

    ws6 = wb.create_sheet('Petunjuk')
    build_petunjuk(ws6)

    output = 'TestExcel_Payroll_HRD.xlsx'
    wb.save(output)
    print(f'Done: {output}')
    print(f'Sheets: {wb.sheetnames}')


if __name__ == '__main__':
    main()
